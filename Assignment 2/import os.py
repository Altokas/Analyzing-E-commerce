import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def export_to_excel(dataframes_dict, filename):
    os.makedirs("exports", exist_ok=True)
    filepath = os.path.join("exports", filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)

    for sheet_name in dataframes_dict.keys():
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        for col in ws.iter_cols(min_row=2, min_col=2):
            if all(isinstance(cell.value, (int, float)) for cell in col if cell.value is not None):
                max_row = ws.max_row
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{max_row}", rule)

    wb.save(filepath)

    total_sheets = len(dataframes_dict)
    total_rows = sum(len(df) for df in dataframes_dict.values())
    print(f"✅ Created file {filename}, {total_sheets} sheets, {total_rows} rows")

# Example usage (replace with your own dataframes):
if __name__ == "__main__":
    df1 = pd.DataFrame({"State": ["SP", "RJ", "MG"], "Orders": [100, 80, 60]})
    df2 = pd.DataFrame({"Payment Type": ["credit_card", "boleto"], "Count": [250, 150]})
    
    export_to_excel(
        {"Orders by State": df1, "Payment Methods": df2},
        "report.xlsx"
    )
